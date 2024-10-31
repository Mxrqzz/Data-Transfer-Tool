from openpyxl import load_workbook

# Carregando as Planilhas
planilhaOrigem = load_workbook('planilhas\dadosOrigem.xlsx')
planilhaDestino = load_workbook('planilhas\dadosDestino.xlsx')

origem = planilhaOrigem.active
destino = planilhaDestino.active

# Dicionario para mapear as lis em cada linha
li_origem = {}
li_destino = {}

# Mapeando as LIs na Planilha de origem
for rowOrigem in range(2, origem.max_row + 1):
    li1 = origem.cell(row=rowOrigem, column=1).value
    if li1 not in li_origem:
        li_origem[li1] = []
    li_origem[li1].append(rowOrigem)

print("Lis Origem ok")

# Mapeando as LIs na Planilha de destino
for rowDestino in range(3, destino.max_row + 1):
    li2 = destino.cell(row=rowDestino, column=2).value
    if li2 not in li_destino:
        li_destino[li2] = []
    li_destino[li2].append(rowDestino)
    
print("Lis Destino ok")

# Copiando o as lis que são iguais
for li_iguais in set(li_origem.keys()).intersection(li_destino.keys()):
    rowsOrigem = li_origem[li_iguais]
    rowsDestino = li_destino[li_iguais]
    
    for rowOrigem, rowDestino in zip(rowsOrigem, rowsDestino):
        #Copiar os dados da colunas B, C, D, E, F e G do arquivo origem para I, J, K,L, M, e N
        for colunaOrigem, colunaDestino in zip(range(2, 8), range(9, 15)):
            valorOrigem = origem.cell(row=rowOrigem, column=colunaOrigem).value
            destino.cell(row=rowDestino, column=colunaDestino).value = valorOrigem
            
print("Dados copiados para arquivo de destino")
        
planilhaDestino.save("planilhas\DadosFinal.xlsx")

print("finalizado")